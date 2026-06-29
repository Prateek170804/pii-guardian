"""Cell-level reversible encryption for real data files (CSV / XLSX).

Companion to the schema-discovery pipeline. Where ``run.py`` classifies *metadata*
and emits Snowflake SQL, this module operates on an actual data file: it reuses the
project's PII detection (``Classifier`` over the same ``cde_dictionary.yaml`` +
value detectors) to decide which columns are sensitive, then encrypts only those
cells with authenticated symmetric crypto (Fernet / AES-128-CBC + HMAC).

Design rules (consistent with CLAUDE.md):
- Connection-free: reads files, writes files. Nothing is sent anywhere.
- Never log or persist a raw sensitive value. Plans, console output and manifest
  carry only column metadata, category, confidence and counts. (Decryption restores
  cleartext to its output file — that is the authorized operation, not a log.)
- Auto-encrypt only high-confidence columns. Borderline columns are surfaced for a
  human and left unselected by default. Lean toward recall when reporting.
- Encryption is idempotent: an already-encrypted cell (``ENC:`` marker) is left as is,
  so re-running never double-encrypts.

This module is the single source of truth for both the CLI (protect.py) and the
web UI (ui.py): build a plan, then encrypt an explicit set of columns.
"""
from __future__ import annotations

import os

import yaml
from cryptography.fernet import Fernet

from .ingest import Column
from .classify import Classifier

# Marker prefix stamped on every encrypted cell. Makes encrypted cells
# self-describing, decryption targeted, and re-encryption idempotent.
MARKER = "ENC:"

# How many non-empty cells per column to sample when running value detection.
_SAMPLE_CAP = 500

# ---------------------------------------------------------------------------
# Config / classifier
# ---------------------------------------------------------------------------
def _load_yaml(path: str) -> dict:
    with open(path, encoding="utf-8") as f:
        return yaml.safe_load(f)


def make_classifier(config_dir: str) -> Classifier:
    """Build a Classifier from the project's config directory."""
    cde = _load_yaml(os.path.join(config_dir, "cde_dictionary.yaml"))
    taxonomy = _load_yaml(os.path.join(config_dir, "taxonomy.yaml"))
    masking_rules = _load_yaml(os.path.join(config_dir, "masking_rules.yaml"))
    return Classifier(cde, taxonomy, masking_rules)


# ---------------------------------------------------------------------------
# File IO helpers (shared by CLI + UI)
# ---------------------------------------------------------------------------
def file_kind(path: str) -> str:
    ext = os.path.splitext(path)[1].lower()
    if ext not in (".csv", ".xlsx"):
        raise ValueError(f"unsupported file type '{ext}'. Use .csv or .xlsx")
    return ext


def _read_csv(path):
    import csv
    with open(path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    if not rows:
        return [], []
    return rows[0], rows[1:]


def _write_csv(path, headers, rows):
    import csv
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(rows)


def _column_values(headers, rows):
    cols = [[] for _ in headers]
    for row in rows:
        for i in range(min(len(row), len(headers))):
            v = (row[i] or "").strip()
            if v:
                cols[i].append(v)
    return cols


def _xlsx_headers_and_values(ws):
    headers = [c.value for c in ws[1]] if ws.max_row >= 1 else []
    cols = [[] for _ in headers]
    for row in ws.iter_rows(min_row=2):
        for i in range(min(len(row), len(headers))):
            v = row[i].value
            if v is not None and str(v).strip():
                cols[i].append(str(v).strip())
    return headers, cols


# ---------------------------------------------------------------------------
# Detection plan
# ---------------------------------------------------------------------------
def _recommend(decision) -> bool:
    """Pre-check ONLY high-confidence (auto) columns. Borderline 'review' columns
    are shown but left UNCHECKED for a human to decide — low-confidence data is
    never auto-applied (CLAUDE.md rule 3). The user can tick review columns by hand
    or use 'Select all flagged'."""
    return decision == "auto"


def _plan_scope(scope, headers, column_values, clf):
    """Plan one scope (a CSV, or one XLSX sheet). One entry per column."""
    plan = []
    for idx, raw_name in enumerate(headers):
        header = str(raw_name).strip() if raw_name is not None else ""
        if not header:
            plan.append({"scope": scope, "index": idx, "name": "", "category": None,
                         "sensitivity": None, "regulations": [], "confidence": 0.0,
                         "plan": "skip", "name_strength": None, "value_detector": None,
                         "value_ratio": None, "recommend": False})
            continue
        col = Column(database="FILE", schema=scope, table="DATA",
                     column=header, data_type="", type_group="OTHER")
        samples = column_values[idx][:_SAMPLE_CAP] if idx < len(column_values) else []
        res = clf.classify(col, samples or None)
        nstr = res.signals.get("name", {}).get("strength")
        vdet = res.signals.get("value", {}).get("detector")
        vratio = res.signals.get("value", {}).get("ratio")
        if res.category and res.confidence >= clf.high:
            decision = "auto"
        elif res.category and res.confidence >= clf.low:
            decision = "review"
        else:
            decision = "skip"
        # Value-inspection guard: a column whose sampled values are effectively
        # constant (<=1 distinct value over enough rows) carries no per-record
        # information and cannot be PII. Suppress weak/ambiguous matches on such
        # columns (e.g. a constant flag/code that merely contains a dictionary word
        # like "acct"). Strong, high-confidence name matches are left untouched.
        if (decision != "skip" and nstr != "strong" and res.confidence < clf.high
                and len(samples) >= 5 and len(set(samples)) <= 1):
            decision = "skip"
        skipped = decision == "skip"
        plan.append({"scope": scope, "index": idx, "name": header,
                     "category": None if skipped else res.category,
                     "sensitivity": None if skipped else res.sensitivity,
                     "regulations": [] if skipped else clf.regulations(res.category, vdet, header),
                     "confidence": res.confidence, "plan": decision, "name_strength": nstr,
                     "value_detector": vdet, "value_ratio": vratio,
                     "recommend": _recommend(decision)})
    return plan


def build_plan(path: str, clf: Classifier):
    """Return a uniform plan (list of column dicts) for a CSV or XLSX file.

    Each entry: scope, index, name, category, confidence, plan (auto|review|skip),
    name_strength, value_detector, value_ratio, recommend.
    """
    if file_kind(path) == ".csv":
        headers, rows = _read_csv(path)
        return _plan_scope("csv", headers, _column_values(headers, rows), clf)
    from openpyxl import load_workbook
    plan = []
    wb = load_workbook(path)
    for ws in wb.worksheets:
        if ws.max_row < 1:
            continue
        headers, colvals = _xlsx_headers_and_values(ws)
        plan += _plan_scope(ws.title, headers, colvals, clf)
    return plan


# ---------------------------------------------------------------------------
# Key handling (generated key file)
# ---------------------------------------------------------------------------
def load_or_create_key(key_path: str):
    """Return (key_bytes, created_bool); generates + writes one if absent."""
    if os.path.exists(key_path):
        with open(key_path, "rb") as f:
            return f.read().strip(), False
    key = Fernet.generate_key()
    with open(key_path, "wb") as f:
        f.write(key)
    try:  # best-effort tighten perms (POSIX); harmless on Windows
        os.chmod(key_path, 0o600)
    except OSError:
        pass
    return key, True


def load_key(key_path: str) -> bytes:
    with open(key_path, "rb") as f:
        return f.read().strip()


# ---------------------------------------------------------------------------
# Cell crypto
# ---------------------------------------------------------------------------
def encrypt_value(fernet: Fernet, value):
    """Encrypt a single cell. Blanks and already-encrypted cells pass through."""
    if value is None:
        return None
    s = str(value)
    if s == "" or s.startswith(MARKER):
        return value
    token = fernet.encrypt(s.encode("utf-8")).decode("ascii")
    return MARKER + token


def decrypt_value(fernet: Fernet, value):
    """Decrypt a single cell if it carries the marker; otherwise pass through."""
    if value is None:
        return None
    s = str(value)
    if not s.startswith(MARKER):
        return value
    return fernet.decrypt(s[len(MARKER):].encode("ascii")).decode("utf-8")


def is_encrypted(value) -> bool:
    return value is not None and str(value).startswith(MARKER)


# ---------------------------------------------------------------------------
# File-level encrypt / decrypt with an explicit column selection
# ---------------------------------------------------------------------------
def encrypt_file(in_path, out_path, fernet, selected, plan=None):
    """Encrypt cells of the selected columns.

    selected : iterable of (scope, column_name) tuples.
    plan     : optional plan list, used only to enrich the returned metadata.
    Returns a list of metadata dicts (scope, column, category, confidence,
    decision, encrypted_cells) — never any raw values.
    """
    selected = {(s, n) for s, n in selected}
    lookup = {(p["scope"], p["name"]): p for p in (plan or [])}
    counts: dict[tuple, int] = {}

    if file_kind(in_path) == ".csv":
        headers, rows = _read_csv(in_path)
        targets = [i for i, h in enumerate(headers) if ("csv", str(h).strip()) in selected]
        for row in rows:
            for i in targets:
                if i < len(row) and row[i] not in (None, "") and not is_encrypted(row[i]):
                    row[i] = encrypt_value(fernet, row[i])
                    counts[("csv", headers[i])] = counts.get(("csv", headers[i]), 0) + 1
        _write_csv(out_path, headers, rows)
    else:
        from openpyxl import load_workbook
        wb = load_workbook(in_path)
        for ws in wb.worksheets:
            if ws.max_row < 1:
                continue
            headers = [c.value for c in ws[1]]
            targets = [i for i, h in enumerate(headers)
                       if (ws.title, str(h).strip() if h is not None else "") in selected]
            for row in ws.iter_rows(min_row=2):
                for i in targets:
                    if i < len(row):
                        cell = row[i]
                        if cell.value in (None, "") or is_encrypted(cell.value):
                            continue
                        cell.value = encrypt_value(fernet, cell.value)
                        key = (ws.title, str(headers[i]).strip())
                        counts[key] = counts.get(key, 0) + 1
        wb.save(out_path)

    entries = []
    for (scope, name) in sorted(selected):
        p = lookup.get((scope, name), {})
        entries.append({"scope": scope, "column": name,
                        "category": p.get("category"), "sensitivity": p.get("sensitivity"),
                        "regulations": p.get("regulations", []),
                        "confidence": p.get("confidence"), "decision": p.get("plan"),
                        "encrypted_cells": counts.get((scope, name), 0)})
    return entries


def decrypt_file(in_path, out_path, fernet) -> int:
    """Decrypt every marked cell in the file. Returns the count decrypted."""
    n = 0
    if file_kind(in_path) == ".csv":
        headers, rows = _read_csv(in_path)
        for row in rows:
            for i in range(len(row)):
                if is_encrypted(row[i]):
                    row[i] = decrypt_value(fernet, row[i])
                    n += 1
        _write_csv(out_path, headers, rows)
    else:
        from openpyxl import load_workbook
        wb = load_workbook(in_path)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if is_encrypted(cell.value):
                        cell.value = decrypt_value(fernet, cell.value)
                        n += 1
        wb.save(out_path)
    return n
